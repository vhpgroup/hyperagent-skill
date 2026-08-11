#!/usr/bin/env python3
"""HSMT Analyzer — GĐ4: Xuất báo cáo Excel 6 tab từ JSON bóc tách + JSON kết quả.

Usage:
  python3 hsmt_excel.py extraction.json [-r results.json] [-o output.xlsx]

results.json (do agent tổng hợp trong quá trình phân tích) — mọi khóa đều TÙY CHỌN:
{
  "meta": {"analysis_date": "..", "coverage_note": ".."},
  "items": {"1": {"candidate": "..", "dat": 10, "khong_dat": 0, "xac_minh": 0,
                   "status": "✅ ĐẠT 100%", "note": ".."}, ...},
  "spec_rows": [["Hạng mục", "Yêu cầu HSMT", "Model", "Giá trị thực", "Trạng thái", "Bằng chứng", "Nguồn"], ...],
  "alternatives": [["Hạng mục (SL)", "Phương án", "Model", "Lưu ý", "Nguồn", "Giá"], ...],
  "analysis_sections": [["A. TIÊU ĐỀ", ["bullet 1", "bullet 2"]], ...],
  "summary": [["Mục", "Nội dung"], ...],
  "traced_models": [["Hạng mục", "Model gốc", "Dấu vân tay", "Độ tin cậy"], ...]
}
Tab 1-2 tự sinh từ extraction.json; tab 3-6 render từ results.json (trống → tab khung).
"""
import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = "Arial"
H_FILL = PatternFill("solid", start_color="1F4E79")
OK_FILL = PatternFill("solid", start_color="E2EFDA")
WARN_FILL = PatternFill("solid", start_color="FFF2CC")
BAD_FILL = PatternFill("solid", start_color="FCE4E4")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=F, bold=True, color="FFFFFF", size=11)
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def add_table(ws, start_row, headers, rows, widths, status_col=None):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, r in enumerate(rows, start_row + 1):
        for j, v in enumerate(r, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(name=F, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN
        if status_col and len(r) >= status_col:
            sv = str(r[status_col - 1])
            fill = OK_FILL if ("Đạt" in sv or "ĐẠT" in sv or "✅" in sv or "🟢" in sv or "🎯" in sv) else (
                BAD_FILL if ("Không đạt" in sv or "KHÔNG ĐẠT" in sv or "🔴" in sv or "LỆCH" in sv) else
                (WARN_FILL if ("xác minh" in sv.lower() or "⚠" in sv) else None))
            if fill:
                for j in range(1, len(headers) + 1):
                    ws.cell(row=i, column=j).fill = fill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return start_row + len(rows) + 1


def title(ws, text, row=1):
    ws.cell(row=row, column=1, value=text).font = Font(name=F, bold=True, size=14, color="1F4E79")
    return row + 2


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("extraction")
    ap.add_argument("-r", "--results", default=None)
    ap.add_argument("-o", "--output", default="HSMT_KetQua_PhanTich.xlsx")
    a = ap.parse_args()

    me = json.load(open(a.extraction, encoding="utf-8"))
    res = json.load(open(a.results, encoding="utf-8")) if a.results else {}
    r_items = res.get("items", {})
    meta = res.get("meta", {})

    wb = Workbook()

    # Tab 1
    ws = wb.active
    ws.title = "1. Thông tin dự án"
    r = title(ws, "THÔNG TIN DỰ ÁN / GÓI THẦU")
    pi = me.get("project_info", {})
    n_req = sum(len(c["requirements"]) for it in me["items"] for c in it["components"])
    n_comp = sum(len(it["components"]) for it in me["items"])
    info = [("Chủ đầu tư", pi.get("investor")), ("Địa chỉ", pi.get("address")),
            ("Tên gói thầu", pi.get("package_name")), ("Nguồn vốn", pi.get("funding_source")),
            ("Phương thức LCNT", pi.get("procurement_method")), ("Loại hợp đồng", pi.get("contract_type")),
            ("Địa điểm thực hiện", pi.get("location")), ("Thời gian thực hiện", pi.get("implementation_time")),
            ("", ""),
            ("Ngày phân tích", meta.get("analysis_date", "")),
            ("Quy mô bóc tách", f"{len(me['items'])} hạng mục | {n_comp} components | {n_req} requirements"),
            ("Ghi chú độ phủ", meta.get("coverage_note", ""))]
    r = add_table(ws, r, ["Trường", "Nội dung"], info, [26, 110])
    r += 1
    ws.cell(row=r, column=1, value="YÊU CẦU CHUNG CỦA HSMT").font = Font(name=F, bold=True, size=12, color="1F4E79")
    r += 1
    add_table(ws, r, ["#", "Nội dung yêu cầu chung"],
              [[i + 1, t] for i, t in enumerate(me.get("general_requirements", []))], [5, 130])
    ws.freeze_panes = "A4"

    # Tab 2
    ws = wb.create_sheet("2. Kết quả phân tích")
    r = title(ws, f"KẾT QUẢ PHÂN TÍCH {len(me['items'])} HẠNG MỤC")
    rows2 = []
    for it in me["items"]:
        n = str(it["item_no"])
        ri = r_items.get(n, {})
        nreq = sum(len(c["requirements"]) for c in it["components"])
        rows2.append([it["item_no"], it["item_name"], it["quantity"], it["unit"], it["match_priority"],
                      len(it["components"]), nreq, ri.get("candidate", ""),
                      ri.get("dat"), ri.get("khong_dat"), ri.get("xac_minh"),
                      ri.get("status", ""), ri.get("note", "")])
    add_table(ws, r, ["STT", "Hạng mục", "SL", "Đơn vị", "Ưu tiên", "Comp", "Req",
                      "Ứng viên chính đề xuất", "Đạt", "Không đạt", "Cần xác minh", "Trạng thái", "Ghi chú"],
              rows2, [5, 28, 7, 9, 8, 6, 6, 42, 6, 9, 11, 22, 48], status_col=12)
    ws.freeze_panes = "A4"

    # Tab 3
    ws = wb.create_sheet("3. So sánh thông số")
    r = title(ws, "SO SÁNH THÔNG SỐ CHI TIẾT — HSMT vs MODEL (bằng chứng trích nguyên văn)")
    add_table(ws, r, ["Hạng mục", "Yêu cầu HSMT (nguyên văn)", "Model đối chiếu", "Giá trị thực tế",
                      "Trạng thái", "Bằng chứng (trích)", "Nguồn [cấp]"],
              res.get("spec_rows", []), [15, 40, 26, 32, 16, 52, 30], status_col=5)
    ws.freeze_panes = "A4"

    # Tab 4
    ws = wb.create_sheet("4. Model tương đương")
    r = title(ws, "DANH MỤC CHÀO THẦU — MODEL CHÍNH + PHƯƠNG ÁN TƯƠNG ĐƯƠNG")
    add_table(ws, r, ["Hạng mục (SL)", "Phương án", "Model", "Lưu ý kỹ thuật", "Nguồn / Datasheet", "Giá tham khảo"],
              res.get("alternatives", []), [22, 13, 48, 46, 34, 24])
    ws.freeze_panes = "A4"

    # Tab 5
    ws = wb.create_sheet("5. Phân tích & Đề xuất")
    r = title(ws, "PHÂN TÍCH CHI TIẾT & ĐỀ XUẤT")
    for sec, bullets in res.get("analysis_sections", []):
        ws.cell(row=r, column=1, value=sec).font = Font(name=F, bold=True, size=12, color="1F4E79")
        r += 1
        for b in bullets:
            c = ws.cell(row=r, column=1, value="• " + b)
            c.font = Font(name=F, size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            ws.row_dimensions[r].height = max(15, 13 * (len(b) // 110 + 1))
            r += 1
        r += 1
    for j in range(1, 9):
        ws.column_dimensions[get_column_letter(j)].width = 16

    # Tab 6
    ws = wb.create_sheet("6. Báo cáo tổng hợp")
    r = title(ws, "BÁO CÁO TỔNG HỢP")
    r = add_table(ws, r, ["Mục", "Nội dung"], res.get("summary", []), [22, 120])
    traced = res.get("traced_models", [])
    if traced:
        r += 1
        ws.cell(row=r, column=1, value=f"BẢNG MODEL GỐC ĐÃ TRUY VẾT ({len(traced)} hạng mục)").font = \
            Font(name=F, bold=True, size=12, color="1F4E79")
        r += 1
        add_table(ws, r, ["Hạng mục", "Model gốc", "Dấu vân tay quyết định", "Độ tin cậy"],
                  traced, [18, 42, 60, 22])
    ws.freeze_panes = "A4"

    # Tab 7 (tùy chọn) — Chứng từ & hồ sơ nhà thầu cần chuẩn bị (bóc từ Chương III + Chương V)
    docs = res.get("documents", {})
    if docs.get("rows"):
        ws = wb.create_sheet("7. Chứng từ & hồ sơ")
        r = title(ws, docs.get("title", "DANH MỤC CHỨNG TỪ & HỒ SƠ NHÀ THẦU CẦN CHUẨN BỊ"))
        if docs.get("note"):
            ws.cell(row=2, column=1, value=docs["note"]).font = Font(name=F, italic=True, size=10)
            r = 4
        r = add_table(ws, r, ["Nhóm", "Chứng từ / Hồ sơ cần làm", "Căn cứ trong E-HSMT",
                              "Phạm vi áp dụng", "Bắt buộc", "Ghi chú chuẩn bị"],
                      docs["rows"], [17, 55, 26, 30, 14, 48], status_col=5)
        if docs.get("note_bottom"):
            ws.cell(row=r + 1, column=1, value=docs["note_bottom"]).font = \
                Font(name=F, bold=True, size=10, color="B3261E")
        ws.freeze_panes = "A5" if docs.get("note") else "A4"

    wb.save(a.output)
    print(f"OK {a.output}: {wb.sheetnames}")


if __name__ == "__main__":
    main()
