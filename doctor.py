#!/usr/bin/env python3
"""
Kiểm tra bundle Hyperagent Document Skills bằng cách CHẠY THẬT.

    python doctor.py

Khác với việc chỉ `import` thư viện: script này tạo file thật, chạy đúng
những script mà SKILL.md bảo agent chạy, rồi đọc kết quả ngược lại. Mục đích
là để bạn biết chắc đường nào dùng được trước khi giao việc thật cho agent.

Exit code 0 nếu mọi thứ bắt buộc đều đạt (phần tuỳ chọn thiếu vẫn coi là đạt).
"""
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
SKILLS = os.path.join(HERE, "skills")

G, R, Y, B, D, X = "\033[32m", "\033[31m", "\033[33m", "\033[34m", "\033[2m", "\033[0m"
if not sys.stdout.isatty() or os.environ.get("NO_COLOR"):
    G = R = Y = B = D = X = ""

results = []          # (category, label, status, detail)  status: ok|fail|skip|warn


def record(cat, label, status, detail=""):
    results.append((cat, label, status, detail))
    icon = {"ok": G + "  OK  " + X, "fail": R + " FAIL " + X,
            "skip": D + " SKIP " + X, "warn": Y + " WARN " + X}[status]
    line = "  [%s] %s" % (icon, label)
    if detail:
        line += D + "  — " + detail + X
    print(line)


def run(cmd, cwd=None, timeout=120):
    """Chạy lệnh, trả về (rc, output gộp stdout+stderr)."""
    try:
        p = subprocess.run(cmd, cwd=cwd, timeout=timeout,
                           stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        return p.returncode, p.stdout.decode("utf-8", "replace").strip()
    except FileNotFoundError:
        return 127, "không tìm thấy lệnh"
    except subprocess.TimeoutExpired:
        return 124, "quá thời gian chờ"


def head(title):
    print("\n" + B + title + X)


# ══════════════════════════════════════════════════ 1. Python
def check_python():
    head("1. Python")
    v = sys.version_info
    vs = "%d.%d.%d" % (v.major, v.minor, v.micro)
    if v >= (3, 10):
        record("python", "Phiên bản Python " + vs, "ok")
    else:
        record("python", "Phiên bản Python " + vs, "fail",
               "cần >= 3.10 (validate.py dùng match/case)")

    required = ["pypdf", "pdfplumber", "openpyxl", "lxml", "defusedxml", "PIL"]
    optional = {"pdf2image": "render PDF ra ảnh",
                "reportlab": "tạo PDF từ đầu",
                "pandas": "phân tích bảng",
                "markitdown": "trích Office sang markdown (python -m markitdown)",
                "pytesseract": "OCR"}
    import importlib
    for m in required:
        try:
            importlib.import_module(m)
            record("python", "import %s" % m, "ok")
        except Exception as e:
            record("python", "import %s" % m, "fail", type(e).__name__)
    for m, why in optional.items():
        try:
            importlib.import_module(m)
            record("python", "import %s" % m, "ok", why)
        except Exception:
            record("python", "import %s" % m, "skip", "tuỳ chọn — mất: " + why)


# ══════════════════════════════════════════════════ 2. Binary hệ thống
def check_binaries():
    head("2. Binary hệ thống")

    rc, out = run(["pdftoppm", "-v"])
    if rc in (0, 99) or "poppler" in out.lower():
        ver = out.splitlines()[0] if out else ""
        record("bin", "poppler (pdftoppm)", "ok", ver[:40])
    else:
        record("bin", "poppler (pdftoppm)", "fail",
               "pdf2image không chạy → mất render PDF ra ảnh")

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        rc, out = run([soffice, "--version"], timeout=90)
        record("bin", "LibreOffice", "ok" if rc == 0 else "warn",
               (out.splitlines()[0][:40] if out else soffice))
    else:
        record("bin", "LibreOffice", "skip",
               "tuỳ chọn — mất: recalc công thức Excel, export PDF, đọc .doc")

    if shutil.which("magick") or shutil.which("convert"):
        record("bin", "ImageMagick", "ok")
    else:
        record("bin", "ImageMagick", "skip",
               "tuỳ chọn — mất: bước crop zoom khi điền PDF form (forms.md)")

    for name, why in (("pandoc", "trích docx sang markdown kèm tracked changes"),
                      ("tesseract", "OCR PDF scan")):
        if shutil.which(name):
            record("bin", name, "ok")
        else:
            record("bin", name, "skip", "tuỳ chọn — mất: " + why)

    if shutil.which("node"):
        rc, out = run(["node", "--version"])
        record("bin", "Node.js", "ok", out)
        np = os.path.join(HERE, "node_modules")
        for pkg in ("docx", "pptxgenjs"):
            if os.path.isdir(os.path.join(np, pkg)):
                record("bin", "npm: " + pkg, "ok")
            else:
                record("bin", "npm: " + pkg, "skip",
                       "tuỳ chọn — mất đường TẠO MỚI file")
    else:
        record("bin", "Node.js", "skip",
               "tuỳ chọn — mất đường TẠO MỚI docx/pptx")


# ══════════════════════════════════════════════════ 3. Chạy thật
def check_xlsx(tmp):
    head("3. xlsx — tạo, giải nén, validate, đóng gói lại")
    S = os.path.join(SKILLS, "xlsx", "scripts")
    src = os.path.join(tmp, "test.xlsx")

    try:
        from openpyxl import Workbook, load_workbook
        wb = Workbook(); ws = wb.active
        ws["A1"] = "Item"; ws["B1"] = "Qty"
        ws["A2"] = "Widget"; ws["B2"] = 5
        ws["B3"] = "=B2*2"
        wb.save(src)
        record("xlsx", "tạo file bằng openpyxl", "ok", "3 dòng, 1 công thức")
    except Exception as e:
        record("xlsx", "tạo file bằng openpyxl", "fail", str(e)[:60])
        return

    unpacked = os.path.join(tmp, "x_unpacked")
    rc, out = run([sys.executable, os.path.join(S, "office", "unpack.py"), src, unpacked])
    if rc == 0:
        n = sum(len(f) for _, _, f in os.walk(unpacked))
        record("xlsx", "office/unpack.py", "ok", "%d file XML" % n)
    else:
        record("xlsx", "office/unpack.py", "fail", out.splitlines()[-1][:70] if out else "")
        return

    # validate.py CỐ Ý chỉ hỗ trợ .docx và .pptx (xem match/case trong
    # office/validate.py). Với .xlsx nó thoát mã 1 — đó là hành vi đúng,
    # không phải lỗi.
    record("xlsx", "office/validate.py", "skip",
           "không áp dụng — validate chỉ hỗ trợ docx/pptx")

    repacked = os.path.join(tmp, "repacked.xlsx")
    rc, out = run([sys.executable, os.path.join(S, "office", "pack.py"),
                   unpacked, repacked, "--original", src])
    if rc != 0 or not os.path.exists(repacked):
        record("xlsx", "office/pack.py", "fail", out.splitlines()[-1][:70] if out else "")
        return
    record("xlsx", "office/pack.py", "ok")

    try:
        from openpyxl import load_workbook
        ws = load_workbook(repacked).active
        good = ws["A2"].value == "Widget" and ws["B2"].value == 5
        record("xlsx", "đọc lại sau roundtrip", "ok" if good else "fail",
               "A2=%r B2=%r" % (ws["A2"].value, ws["B2"].value))
    except Exception as e:
        record("xlsx", "đọc lại sau roundtrip", "fail", str(e)[:60])

    recalc = os.path.join(S, "recalc.py")
    if shutil.which("soffice") or shutil.which("libreoffice"):
        rc, out = run([sys.executable, recalc, src], cwd=S, timeout=180)
        record("xlsx", "recalc.py (cần LibreOffice)", "ok" if rc == 0 else "warn",
               (out.splitlines()[-1][:70] if out else ""))
    else:
        record("xlsx", "recalc.py", "skip", "không có LibreOffice")


def check_pdf(tmp):
    head("4. pdf — tạo, trích text, render ảnh")
    S = os.path.join(SKILLS, "pdf", "scripts")
    src = os.path.join(tmp, "test.pdf")

    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        c = canvas.Canvas(src, pagesize=letter)
        c.drawString(100, 700, "Hyperagent doctor smoke test")
        c.drawString(100, 680, "dong thu hai")
        c.save()
        record("pdf", "tạo PDF bằng reportlab", "ok")
    except Exception as e:
        record("pdf", "tạo PDF bằng reportlab", "skip", "thiếu reportlab")
        return

    try:
        from pypdf import PdfReader
        r = PdfReader(src)
        record("pdf", "pypdf đọc được", "ok", "%d trang" % len(r.pages))
    except Exception as e:
        record("pdf", "pypdf đọc được", "fail", str(e)[:60])

    try:
        import pdfplumber
        with pdfplumber.open(src) as pdf:
            txt = pdf.pages[0].extract_text() or ""
        ok = "doctor" in txt
        record("pdf", "pdfplumber trích text", "ok" if ok else "warn",
               repr(txt[:35]))
    except Exception as e:
        record("pdf", "pdfplumber trích text", "fail", str(e)[:60])

    rc, out = run([sys.executable, os.path.join(S, "check_fillable_fields.py"), src])
    record("pdf", "check_fillable_fields.py", "ok" if rc == 0 else "fail",
           (out.splitlines()[-1][:60] if out else ""))

    rc, out = run([sys.executable, os.path.join(S, "extract_form_structure.py"),
                   src, os.path.join(tmp, "structure.json")])
    record("pdf", "extract_form_structure.py", "ok" if rc == 0 else "fail",
           (out.splitlines()[-1][:60] if out else ""))

    # Chú ý: script này KHÔNG tự tạo thư mục đích, phải mkdir trước.
    imgdir = os.path.join(tmp, "pages")
    os.makedirs(imgdir, exist_ok=True)
    rc, out = run([sys.executable, os.path.join(S, "convert_pdf_to_images.py"), src, imgdir])
    if rc == 0 and os.path.isdir(imgdir) and os.listdir(imgdir):
        record("pdf", "convert_pdf_to_images.py (cần poppler)", "ok",
               "%d ảnh" % len(os.listdir(imgdir)))
    else:
        record("pdf", "convert_pdf_to_images.py (cần poppler)", "fail",
               out.splitlines()[-1][:60] if out else "không tạo được ảnh")


# .docx tối thiểu nhưng hợp lệ — dùng làm mẫu khi máy không có LibreOffice.
# Chỉ 3 part là đủ để Word mở được và để DOCXSchemaValidator chạy thật.
_MIN_DOCX = {
    "[Content_Types].xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>""",
    "_rels/.rels": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>""",
    "word/document.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>
<w:p><w:r><w:t>Hyperagent doctor smoke test</w:t></w:r></w:p>
<w:sectPr><w:pgSz w:w="12240" w:h="15840"/></w:sectPr>
</w:body>
</w:document>""",
}


def _write_min_docx(path):
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for name, body in _MIN_DOCX.items():
            z.writestr(name, body)


_PPTX_GEN_JS = """
const pptxgen = require("pptxgenjs");
const pres = new pptxgen();
const s = pres.addSlide();
s.addText("Hyperagent doctor smoke test", { x:0.5, y:1.5, fontSize:28, bold:true });
pres.writeFile({ fileName: process.argv[2] }).then(() => console.log("ok"));
"""


def _make_pptx_with_pptxgenjs(path, tmp):
    """Sinh .pptx bằng pptxgenjs. Trả về True nếu thành công."""
    if not shutil.which("node"):
        return False
    nm = os.path.join(HERE, "node_modules")
    if not os.path.isdir(os.path.join(nm, "pptxgenjs")):
        return False
    js = os.path.join(tmp, "_gen_pptx.js")
    with open(js, "w") as fh:
        fh.write(_PPTX_GEN_JS)
    env = dict(os.environ)
    env["NODE_PATH"] = nm + os.pathsep + env.get("NODE_PATH", "")
    try:
        p = subprocess.run(["node", js, path], cwd=tmp, env=env, timeout=120,
                           stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        return p.returncode == 0 and os.path.exists(path)
    except Exception:
        return False


# pptxgenjs xuất notesMasterIdLst SAU sldIdLst, trái thứ tự mà pml.xsd đòi
# (sldMasterIdLst → notesMasterIdLst → sldIdLst → sldSz). PowerPoint vẫn mở
# được, nhưng validate.py bắt đúng theo schema. Đây là lỗi thượng nguồn của
# pptxgenjs, không phải của skill.
_KNOWN_PPTXGENJS_QUIRK = "notesMasterIdLst"


def check_docx_pptx(tmp):
    head("5. docx / pptx — giải nén, validate, đóng gói lại")
    soffice = shutil.which("soffice") or shutil.which("libreoffice")

    for kind in ("docx", "pptx"):
        S = os.path.join(SKILLS, kind, "scripts")
        src = os.path.join(tmp, "sample." + kind)

        # Cần một file mẫu. Dùng LibreOffice tạo từ file text là cách chắc nhất
        # mà không phụ thuộc Node.
        made = False
        if soffice:
            txt = os.path.join(tmp, "seed_%s.txt" % kind)
            with open(txt, "w") as fh:
                fh.write("Hyperagent doctor smoke test\nDong thu hai\n")
            outdir = os.path.join(tmp, "so_" + kind)
            os.makedirs(outdir, exist_ok=True)
            target = "docx" if kind == "docx" else "pptx:Impress MS PowerPoint 2007 XML"
            rc, out = run([soffice, "--headless", "--convert-to", target,
                           "--outdir", outdir, txt], timeout=180)
            cand = os.path.join(outdir, "seed_%s.%s" % (kind, kind))
            if rc == 0 and os.path.exists(cand):
                shutil.copy(cand, src); made = True

        if not made and kind == "pptx" and _make_pptx_with_pptxgenjs(src, tmp):
            made = True
            record(kind, "tạo file mẫu bằng pptxgenjs", "ok")
        elif made:
            record(kind, "tạo file mẫu bằng LibreOffice", "ok")
        elif kind == "docx":
            # Không có LibreOffice vẫn test được docx: tự dựng OOXML tối thiểu.
            try:
                _write_min_docx(src)
                made = True
                record(kind, "tạo file mẫu (OOXML tối thiểu, không cần LibreOffice)", "ok")
            except Exception as e:
                record(kind, "tạo file mẫu", "fail", str(e)[:60])
        if not made:
            record(kind, "tạo file mẫu", "skip",
                   "cần LibreOffice để sinh file .pptx mẫu — bỏ qua nhóm test này")
            continue

        unpacked = os.path.join(tmp, kind + "_unpacked")
        rc, out = run([sys.executable, os.path.join(S, "office", "unpack.py"), src, unpacked])
        if rc != 0:
            record(kind, "office/unpack.py", "fail", out.splitlines()[-1][:70] if out else "")
            continue
        n = sum(len(f) for _, _, f in os.walk(unpacked))
        record(kind, "office/unpack.py", "ok", "%d file XML" % n)

        rc, out = run([sys.executable, os.path.join(S, "office", "validate.py"), src])
        if rc == 0:
            record(kind, "office/validate.py", "ok",
                   (out.splitlines()[-1][:70] if out else ""))
        elif _KNOWN_PPTXGENJS_QUIRK in out:
            record(kind, "office/validate.py", "warn",
                   "lỗi thứ tự notesMasterIdLst có sẵn trong output của pptxgenjs, "
                   "không phải do skill")
        else:
            record(kind, "office/validate.py", "fail",
                   (out.splitlines()[-1][:70] if out else ""))

        repacked = os.path.join(tmp, "repacked." + kind)
        rc, out = run([sys.executable, os.path.join(S, "office", "pack.py"),
                       unpacked, repacked, "--original", src])
        if rc == 0 and os.path.exists(repacked):
            try:
                zipfile.ZipFile(repacked).testzip()
                record(kind, "office/pack.py + roundtrip", "ok")
            except Exception as e:
                record(kind, "office/pack.py + roundtrip", "fail", str(e)[:60])
        else:
            record(kind, "office/pack.py", "fail", out.splitlines()[-1][:70] if out else "")

        # thumbnail.py: cần LibreOffice cho .pptx, nhưng nhận thẳng .pdf được,
        # nên vẫn test được phần ghép lưới khi máy không có LibreOffice.
        if kind == "pptx":
            thumb = os.path.join(S, "thumbnail.py")
            if not os.path.isfile(thumb):
                record(kind, "thumbnail.py", "fail", "thiếu file")
            else:
                target = src if soffice else os.path.join(tmp, "test.pdf")
                if not os.path.exists(target):
                    record(kind, "thumbnail.py", "skip", "không có file đầu vào")
                else:
                    outjpg = os.path.join(tmp, "thumbnails.jpg")
                    rc, out = run([sys.executable, thumb, target, "-o", outjpg],
                                  timeout=300)
                    if rc == 0 and os.path.exists(outjpg):
                        record(kind, "thumbnail.py", "ok",
                               (out.splitlines()[-1][:60] if out else "") +
                               ("" if soffice else " (qua .pdf, không có LibreOffice)"))
                    else:
                        record(kind, "thumbnail.py", "fail",
                               out.splitlines()[-1][:70] if out else "")


# ══════════════════════════════════════════════════ 6. research / browser / interactive
def check_tool_skills(tmp):
    head("6. research / browser / interactive")

    # ---- research ----
    R = os.path.join(SKILLS, "research", "scripts")
    rc, out = run([sys.executable, os.path.join(R, "fetch.py"),
                   "https://example.com", "--max-chars", "300"], timeout=90)
    if rc == 0 and "Example" in out:
        how = next((l for l in out.splitlines() if l.startswith("Bóc bằng")), "")
        record("research", "fetch.py trên trang thật", "ok", how[:50])
    else:
        record("research", "fetch.py trên trang thật", "warn",
               "không tải được — kiểm tra mạng/proxy")

    if os.environ.get("EXA_API_KEY"):
        rc, out = run([sys.executable, os.path.join(R, "search.py"),
                       "test query", "-n", "2"], timeout=90)
        record("research", "search.py backend exa", "ok" if rc == 0 else "fail",
               (out.splitlines()[0][:60] if out else ""))
    else:
        record("research", "search.py backend exa", "skip",
               "chưa đặt EXA_API_KEY — backend mặc định sẽ báo lỗi")
        rc, out = run([sys.executable, os.path.join(R, "search.py"),
                       "python zipfile", "--backend", "ddg", "-n", "2"], timeout=120)
        record("research", "search.py fallback ddg", "ok" if rc == 0 else "warn",
               "" if rc == 0 else "ddg hay bị rate-limit, không phải lỗi code")

    # ---- browser ----
    B = os.path.join(SKILLS, "browser", "scripts", "browse.py")
    try:
        import playwright  # noqa: F401
        has_pw = True
    except ImportError:
        has_pw = False
    if not has_pw:
        record("browser", "Playwright", "skip",
               "chưa cài: pip install playwright && python -m playwright install chromium")
    else:
        rc, out = run([sys.executable, B, "get", "https://example.com"], timeout=180)
        if rc == 0 and "Example" in out:
            record("browser", "browse.py get (Chromium thật)", "ok")
        else:
            record("browser", "browse.py get", "fail",
                   (out.splitlines()[-1][:70] if out else ""))
        shot = os.path.join(tmp, "shot.png")
        rc, out = run([sys.executable, B, "shot", "https://example.com", "-o", shot],
                      timeout=180)
        record("browser", "browse.py shot", "ok" if os.path.exists(shot) else "fail",
               (out.splitlines()[-1][:60] if out else ""))

    # ---- interactive ----
    I = os.path.join(SKILLS, "interactive", "scripts")
    md = os.path.join(tmp, "deck.md")
    with open(md, "w", encoding="utf-8") as fh:
        fh.write("# Slide một\n\n- điểm a\n- điểm b\n\n---\n\n# Slide hai\n\nNội dung.\n")
    deck = os.path.join(tmp, "deck.html")
    rc, out = run([sys.executable, os.path.join(I, "new_deck.py"), md, "-o", deck],
                  timeout=60)
    if rc == 0 and os.path.exists(deck):
        body = open(deck, encoding="utf-8").read()
        needed = ["ArrowRight", "touchstart", "close-fullscreen", "navigate"]
        miss = [k for k in needed if k not in body]
        record("interactive", "new_deck.py", "ok" if not miss else "fail",
               "đủ điều hướng" if not miss else "thiếu: " + ", ".join(miss))
    else:
        record("interactive", "new_deck.py", "fail",
               (out.splitlines()[-1][:70] if out else ""))
    rc, out = run([sys.executable, "-c",
                   "import ast,sys; ast.parse(open(sys.argv[1],encoding='utf-8').read())",
                   os.path.join(I, "preview.py")], timeout=30)
    record("interactive", "preview.py", "ok" if rc == 0 else "fail")


# ══════════════════════════════════════════════════ tổng kết
def summary():
    head("Tổng kết")
    n_ok = sum(1 for r in results if r[2] == "ok")
    n_fail = sum(1 for r in results if r[2] == "fail")
    n_skip = sum(1 for r in results if r[2] == "skip")
    n_warn = sum(1 for r in results if r[2] == "warn")

    print("  %s%d đạt%s   %s%d hỏng%s   %s%d cảnh báo%s   %s%d bỏ qua%s" %
          (G, n_ok, X, R, n_fail, X, Y, n_warn, X, D, n_skip, X))

    if n_fail:
        print("\n  " + R + "Những mục hỏng:" + X)
        for cat, label, status, detail in results:
            if status == "fail":
                print("    - [%s] %s %s" % (cat, label, D + detail + X))
        print("\n  Skill tương ứng sẽ không chạy đúng. Sửa xong chạy lại doctor.py.")
    else:
        print("\n  " + G + "Mọi thứ bắt buộc đều đạt." + X)

    if n_skip:
        print("\n  " + D + "Mục bỏ qua là tính năng tuỳ chọn — xem cột giải thích"
              " để biết mất gì." + X)
    return 1 if n_fail else 0


def main():
    print()
    print(B + "Hyperagent Document Skills — kiểm tra cài đặt" + X)
    print(D + "Chạy thật, không chỉ kiểm tra import." + X)

    if not os.path.isdir(SKILLS):
        print(R + "\nKhông tìm thấy thư mục skills/ cạnh doctor.py" + X)
        return 2

    tmp = tempfile.mkdtemp(prefix="hyperagent_doctor_")
    try:
        check_python()
        check_binaries()
        check_xlsx(tmp)
        check_pdf(tmp)
        check_docx_pptx(tmp)
        check_tool_skills(tmp)
        return summary()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
        print()


if __name__ == "__main__":
    sys.exit(main())
